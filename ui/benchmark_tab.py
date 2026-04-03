"""벤치마크 탭: 최대 10개 모델 성능 비교"""
import os
import re

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
    QLabel, QPushButton, QSpinBox, QComboBox,
    QTableWidget, QTableWidgetItem, QProgressBar,
    QFileDialog, QHeaderView, QSizePolicy, QMessageBox,
    QLineEdit, QScrollArea,
)

from core.benchmark_runner import BenchmarkConfig, BenchmarkRunner

_MAX_MODELS = 10
_WARMUP = 300

_RESULT_COLS = [
    "모델명", "타입", "Provider",
    "원본 해상도", "모델 입력", "배치", "dtype",
    "워밍업", "반복수",
    "전처리 ms", "추론 ms", "후처리 ms", "총 ms",
    "최소 ms", "최대 ms", "표준편차",
    "FPS (img/s)", "CPU %", "RAM MB", "GPU %", "VRAM MB",
]


def _parse_src_hw(text: str) -> "tuple[int, int]":
    """'1920×1080' (W×H) → (H, W). 빈 문자열이면 기본값 (1080, 1920) 반환."""
    text = text.strip()
    if not text:
        return (1080, 1920)
    nums = [int(n) for n in re.findall(r"\d+", text) if int(n) > 0]
    if len(nums) == 1:
        return (nums[0], nums[0])
    if len(nums) >= 2:
        return (nums[1], nums[0])   # W×H 표기 → (H, W) 내부 표현
    return (1080, 1920)


class ModelSlotWidget(QGroupBox):
    removed = Signal(object)
    extra_models_selected = Signal(list)  # 다중 선택 시 추가 파일 경로들

    def __init__(self, idx: int, parent=None):
        super().__init__(f"모델 {idx + 1}", parent)
        self._path = ""

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 6, 8, 6)
        layout.setSpacing(4)

        # ── 파일 선택 + 타입 + 제거 ─────────────────────────────────────
        row1 = QHBoxLayout()
        row1.setSpacing(6)

        self._lbl_path = QLabel("모델 파일을 선택하세요")
        self._lbl_path.setStyleSheet("color: gray; font-size: 11px;")
        self._lbl_path.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        btn_sel = QPushButton("모델 선택…")
        btn_sel.setFixedWidth(90)
        btn_sel.clicked.connect(self._select)

        self._type_combo = QComboBox()
        self._type_combo.addItems(["YOLO", "CenterNet"])
        self._type_combo.setFixedWidth(86)
        self._type_combo.setToolTip("모델 아키텍처 타입")

        btn_rm = QPushButton("X")
        btn_rm.setFixedSize(24, 24)
        btn_rm.setStyleSheet("color: red; font-weight: bold; border: none;")
        btn_rm.setToolTip("이 슬롯 제거")
        btn_rm.clicked.connect(lambda: self.removed.emit(self))

        row1.addWidget(self._lbl_path)
        row1.addWidget(btn_sel)
        row1.addWidget(self._type_combo)
        row1.addWidget(btn_rm)
        layout.addLayout(row1)

        # ── 배치 크기 + 원본 해상도 ──────────────────────────────────────
        row2 = QHBoxLayout()
        row2.setSpacing(10)

        row2.addWidget(QLabel("배치:"))
        self._batch_spin = QSpinBox()
        self._batch_spin.setRange(1, 64)
        self._batch_spin.setValue(1)
        self._batch_spin.setFixedWidth(55)
        self._batch_spin.setToolTip(
            "입력 배치 크기. 모델이 고정 배치를 요구하면 자동으로 조정됩니다."
        )
        row2.addWidget(self._batch_spin)

        row2.addSpacing(8)
        row2.addWidget(QLabel("원본 해상도:"))
        self._src_hw_edit = QLineEdit()
        self._src_hw_edit.setText("1920×1080")
        self._src_hw_edit.setFixedWidth(120)
        self._src_hw_edit.setToolTip(
            "전처리 전 원본 프레임 해상도 (W×H).\n"
            "모델 입력 크기로 resize된 뒤 추론합니다.\n"
            "예: 1920×1080 / 3840×2160 / 1280×720"
        )
        row2.addWidget(self._src_hw_edit)
        row2.addStretch()
        layout.addLayout(row2)

        # ── EP(Execution Provider) 선택 ──────────────────────────────────
        from core.ep_manager import EP_VARIANTS, is_ep_available
        row3 = QHBoxLayout()
        row3.setSpacing(8)
        row3.addWidget(QLabel("EP:"))
        self._ep_combo = QComboBox()
        self._ep_combo.setFixedWidth(300)
        self._ep_combo.addItem("자동  (현재 환경 기본값)", "auto")
        for ep_key, info in EP_VARIANTS.items():
            label = info["label"]
            avail = is_ep_available(ep_key)
            display = label if avail else f"{label}  [미설치]"
            self._ep_combo.addItem(display, ep_key)
            if not avail:
                # 미설치 항목 비활성화
                idx = self._ep_combo.count() - 1
                item = self._ep_combo.model().item(idx)
                if item:
                    item.setEnabled(False)
        self._ep_combo.setToolTip(
            "실행에 사용할 Execution Provider.\n"
            "미설치 항목은 scripts/setup_ep.bat을 실행하여 설치하세요."
        )
        row3.addWidget(self._ep_combo)
        row3.addStretch()
        layout.addLayout(row3)

    def _select(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "ONNX 모델 선택", "", "ONNX 모델 (*.onnx)"
        )
        if not paths:
            return
        # 첫 번째 파일은 이 슬롯에 설정
        self._path = paths[0]
        self._lbl_path.setText(os.path.basename(paths[0]))
        self._lbl_path.setStyleSheet("font-size: 11px;")
        self._lbl_path.setToolTip(paths[0])
        # 나머지 파일은 시그널로 부모에게 전달
        if len(paths) > 1:
            self.extra_models_selected.emit(paths[1:])

    def get_config(self, iterations: int) -> "BenchmarkConfig | None":
        if not self._path:
            return None
        model_type = "yolo" if self._type_combo.currentIndex() == 0 else "darknet"
        src_hw = _parse_src_hw(self._src_hw_edit.text())
        ep_key = self._ep_combo.currentData()
        return BenchmarkConfig(
            model_path=self._path,
            model_type=model_type,
            iterations=iterations,
            warmup=_WARMUP,
            batch_size=self._batch_spin.value(),
            src_hw=src_hw,
            ep_key=ep_key,
        )


class BenchmarkTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._runner: "BenchmarkRunner | None" = None
        self._slots: list[ModelSlotWidget] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(8)

        # ── 모델 설정 그룹 (최대 4슬롯 높이까지 스크롤 없이 표시) ────────
        self._model_group = QGroupBox("모델 설정")
        self._model_layout = QVBoxLayout(self._model_group)
        self._model_layout.setSpacing(6)

        self._btn_add = QPushButton("＋   모델 추가  (최대 10개)")
        self._btn_add.setFixedHeight(28)
        self._btn_add.clicked.connect(self._add_slot)
        self._model_layout.addWidget(self._btn_add)

        scroll = QScrollArea()
        scroll.setWidget(self._model_group)
        scroll.setWidgetResizable(True)
        # 4슬롯 높이(~560px)까지는 스크롤 없이 표시, 이후 스크롤
        scroll.setMinimumHeight(560)
        scroll.setMaximumHeight(560)
        root.addWidget(scroll)

        self._add_slot()   # 기본 슬롯 1개

        # ── 벤치마크 설정 ───────────────────────────────────────────────
        cfg_group = QGroupBox("벤치마크 설정")
        cfg_row = QHBoxLayout(cfg_group)
        cfg_row.setSpacing(12)

        cfg_row.addWidget(QLabel("워밍업:"))
        lbl_wu = QLabel(f"{_WARMUP}  (고정)")
        lbl_wu.setStyleSheet("font-weight: bold;")
        cfg_row.addWidget(lbl_wu)

        cfg_row.addSpacing(24)
        cfg_row.addWidget(QLabel("반복 횟수:"))
        self._iter_spin = QSpinBox()
        self._iter_spin.setRange(10, 5000)
        self._iter_spin.setValue(500)
        self._iter_spin.setSingleStep(100)
        self._iter_spin.setFixedWidth(80)
        cfg_row.addWidget(self._iter_spin)
        cfg_row.addStretch()

        root.addWidget(cfg_group)

        # ── 실행 / 중지 ─────────────────────────────────────────────────
        act_row = QHBoxLayout()
        self._btn_run = QPushButton("벤치마크 실행")
        self._btn_run.setFixedHeight(34)
        self._btn_run.setStyleSheet("font-size: 13px; font-weight: bold;")
        self._btn_run.clicked.connect(self._on_run)

        self._btn_stop = QPushButton("중지")
        self._btn_stop.setFixedSize(90, 34)
        self._btn_stop.setEnabled(False)
        self._btn_stop.setStyleSheet("color: black;")
        self._btn_stop.clicked.connect(self._on_stop)

        act_row.addWidget(self._btn_run)
        act_row.addWidget(self._btn_stop)
        root.addLayout(act_row)

        # ── 진행 상태 ───────────────────────────────────────────────────
        self._progress = QProgressBar()
        self._progress.setTextVisible(True)
        self._progress.setValue(0)
        self._status_lbl = QLabel("대기 중")
        self._status_lbl.setStyleSheet("font-size: 11px; color: #666;")
        root.addWidget(self._progress)
        root.addWidget(self._status_lbl)

        # ── 결과 테이블 ─────────────────────────────────────────────────
        result_group = QGroupBox("결과")
        result_vl = QVBoxLayout(result_group)

        tbl_hdr = QHBoxLayout()
        tbl_hdr.addStretch()

        btn_export = QPushButton("Excel 내보내기")
        btn_export.setFixedWidth(110)
        btn_export.clicked.connect(self._on_export)
        tbl_hdr.addWidget(btn_export)

        btn_clear = QPushButton("결과 지우기")
        btn_clear.setFixedWidth(90)
        btn_clear.clicked.connect(lambda: self._table.setRowCount(0))
        tbl_hdr.addWidget(btn_clear)
        result_vl.addLayout(tbl_hdr)

        self._table = QTableWidget(0, len(_RESULT_COLS))
        self._table.setHorizontalHeaderLabels(_RESULT_COLS)
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for i in range(1, len(_RESULT_COLS)):
            self._table.horizontalHeader().setSectionResizeMode(
                i, QHeaderView.ResizeToContents
            )
        self._table.setSelectionBehavior(QTableWidget.SelectRows)
        self._table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._table.verticalHeader().setVisible(False)
        self._table.setAlternatingRowColors(True)
        result_vl.addWidget(self._table)

        root.addWidget(result_group, stretch=1)

    # ── 슬롯 관리 ───────────────────────────────────────────────────────
    def _add_slot(self):
        if len(self._slots) >= _MAX_MODELS:
            return
        slot = ModelSlotWidget(len(self._slots))
        slot.removed.connect(self._remove_slot)
        slot.extra_models_selected.connect(self._on_extra_models)
        self._slots.append(slot)
        insert_at = self._model_layout.indexOf(self._btn_add)
        self._model_layout.insertWidget(insert_at, slot)
        self._btn_add.setEnabled(len(self._slots) < _MAX_MODELS)
        return slot

    def _remove_slot(self, slot: ModelSlotWidget):
        if slot not in self._slots:
            return
        self._slots.remove(slot)
        self._model_layout.removeWidget(slot)
        slot.deleteLater()
        for i, s in enumerate(self._slots):
            s.setTitle(f"모델 {i + 1}")
        self._btn_add.setEnabled(True)

    def _on_extra_models(self, paths: list):
        """다중 선택 시 추가 파일들을 새 슬롯에 자동 배치"""
        for path in paths:
            if len(self._slots) >= _MAX_MODELS:
                break
            slot = self._add_slot()
            if slot:
                slot._path = path
                slot._lbl_path.setText(os.path.basename(path))
                slot._lbl_path.setStyleSheet("font-size: 11px;")
                slot._lbl_path.setToolTip(path)

    # ── 벤치마크 실행 ───────────────────────────────────────────────────
    def _on_run(self):
        if self._runner and self._runner.isRunning():
            return

        configs = [
            cfg for slot in self._slots
            if (cfg := slot.get_config(self._iter_spin.value()))
        ]
        if not configs:
            QMessageBox.information(
                self, "알림",
                "벤치마크할 ONNX 모델을 하나 이상 선택해주세요."
            )
            return

        total = sum(c.warmup + c.iterations for c in configs)
        self._progress.setMaximum(total)
        self._progress.setValue(0)
        self._status_lbl.setText("벤치마크 시작...")
        self._btn_run.setEnabled(False)
        self._btn_stop.setEnabled(True)

        runner = BenchmarkRunner(configs)
        self._runner = runner
        runner.progress_updated.connect(self._on_progress)
        runner.result_ready.connect(self._on_result)
        runner.finished.connect(self._on_finished)
        runner.error.connect(lambda msg: self._status_lbl.setText(f"오류: {msg}"))
        runner.start()

    def _on_stop(self):
        if self._runner:
            self._runner.stop()
        self._btn_stop.setEnabled(False)
        self._status_lbl.setText("중지 요청됨…")

    def _on_progress(self, current: int, total: int, msg: str):
        self._progress.setMaximum(total)
        self._progress.setValue(current)
        self._status_lbl.setText(msg)

    def _on_result(self, result):
        row = self._table.rowCount()
        self._table.insertRow(row)

        vram = (
            f"{result.gpu_mem_used} / {result.gpu_mem_total} MB"
            if result.gpu_mem_used is not None else "N/A"
        )
        values = [
            result.model_name,
            result.model_type.upper(),
            result.provider,
            f"{result.src_size[1]}×{result.src_size[0]}",     # W×H 표시
            f"{result.model_size[1]}×{result.model_size[0]}", # W×H 표시
            str(result.batch_size),
            result.input_dtype,
            str(result.warmup_count),
            str(result.iter_count),
            f"{result.mean_pre_ms:.2f}",
            f"{result.mean_infer_ms:.2f}",
            f"{result.mean_post_ms:.2f}",
            f"{result.mean_total_ms:.2f}",
            f"{result.min_ms:.2f}",
            f"{result.max_ms:.2f}",
            f"{result.std_ms:.2f}",
            f"{result.fps:.1f}",
            f"{result.cpu_pct:.1f}",
            f"{result.ram_mb:.0f}",
            f"{result.gpu_pct}%" if result.gpu_pct is not None else "N/A",
            vram,
        ]
        for col, val in enumerate(values):
            item = QTableWidgetItem(val)
            item.setTextAlignment(Qt.AlignCenter)
            self._table.setItem(row, col, item)
        self._table.scrollToBottom()

    def _on_finished(self):
        self._btn_run.setEnabled(True)
        self._btn_stop.setEnabled(False)
        self._progress.setValue(self._progress.maximum())
        n = self._table.rowCount()
        self._status_lbl.setText(f"완료  —  {n}개 결과")
        self._highlight_best_fps()

    def _highlight_best_fps(self):
        """FPS 열에서 최고값 행을 초록색으로 하이라이트"""
        from PySide6.QtGui import QColor
        fps_col = _RESULT_COLS.index("FPS (img/s)")
        best_row, best_fps = -1, -1.0
        for row in range(self._table.rowCount()):
            item = self._table.item(row, fps_col)
            if item:
                try:
                    val = float(item.text())
                    if val > best_fps:
                        best_fps, best_row = val, row
                except ValueError:
                    pass
        if best_row < 0:
            return
        highlight = QColor(76, 175, 80, 60)  # 초록색 반투명
        for col in range(self._table.columnCount()):
            item = self._table.item(best_row, col)
            if item:
                item.setBackground(highlight)

    # ── Excel 내보내기 ──────────────────────────────────────────────────
    def _on_export(self):
        if self._table.rowCount() == 0:
            QMessageBox.information(self, "알림", "내보낼 결과가 없습니다.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Excel로 내보내기", "benchmark_results.xlsx",
            "Excel 파일 (*.xlsx)"
        )
        if not path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            QMessageBox.critical(
                self, "패키지 없음",
                "openpyxl이 설치되지 않았습니다.\n\npip install openpyxl"
            )
            return

        try:
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.title = "벤치마크 결과"

            # 헤더 행
            headers = [
                self._table.horizontalHeaderItem(c).text()
                for c in range(self._table.columnCount())
            ]
            ws.append(headers)
            header_fill = PatternFill("solid", fgColor="2E4D7B")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 데이터 행 (숫자는 숫자 타입으로 변환)
            for row in range(self._table.rowCount()):
                row_data: list = []
                for col in range(self._table.columnCount()):
                    item = self._table.item(row, col)
                    text = item.text() if item else ""
                    cleaned = (
                        text.replace(" ms", "")
                            .replace(" MB", "")
                            .replace("%", "")
                    )
                    try:
                        val: object = (
                            float(cleaned) if "." in cleaned else int(cleaned)
                        )
                    except ValueError:
                        val = text
                    row_data.append(val)
                ws.append(row_data)
                # 데이터 행 가운데 정렬
                for cell in ws[ws.max_row]:
                    cell.alignment = Alignment(horizontal="center")

            # 컬럼 너비 자동 조정
            for col_cells in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(  # type: ignore[union-attr]
                    max_len + 3, 32
                )

            wb.save(path)
            QMessageBox.information(self, "저장 완료", f"저장 완료:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 실패:\n{e}")

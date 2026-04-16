"""Pseudo Labeling + 정량적 평가 탭"""
import os
import glob

import cv2
import numpy as np
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QPushButton,
    QLabel, QLineEdit, QDoubleSpinBox, QProgressBar, QFileDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox, QTextEdit,
    QComboBox, QTabWidget, QDialog, QCheckBox,
)

from core.model_loader import ModelInfo, load_model
from core.inference import (run_inference, convert_darknet_to_unified,
                            run_classification, ClassificationResult, UNIFIED_NAMES)
from ui import theme


# ------------------------------------------------------------------ #
# Pseudo Labeling 워커
# ------------------------------------------------------------------ #
class _PseudoWorker(QThread):
    progress = Signal(int, int)
    finished_ok = Signal(str)
    error = Signal(str)

    def __init__(self, model_info, src_path, out_dir, conf,
                 source_type="image", frame_interval=1, save_frames=False):
        super().__init__()
        self.model_info = model_info
        self.src_path = src_path
        self.out_dir = out_dir
        self.conf = conf
        self.source_type = source_type      # "image" or "video"
        self.frame_interval = max(1, frame_interval)
        self.save_frames = save_frames

    def _write_label(self, txt_path, res, h, w):
        with open(txt_path, "w") as f:
            for box, score, cid in zip(res.boxes, res.scores, res.class_ids):
                x1, y1, x2, y2 = box
                cx = ((x1 + x2) / 2) / w
                cy = ((y1 + y2) / 2) / h
                bw = (x2 - x1) / w
                bh = (y2 - y1) / h
                f.write(f"{int(cid)} {cx:.6f} {cy:.6f} {bw:.6f} {bh:.6f}\n")

    def _infer(self, frame):
        res = run_inference(self.model_info, frame, self.conf)
        if self.model_info.model_type == "darknet":
            res = convert_darknet_to_unified(res)
        return res

    def _run_images(self):
        exts = ("*.jpg", "*.jpeg", "*.png", "*.bmp")
        files = []
        for e in exts:
            files.extend(glob.glob(os.path.join(self.src_path, e)))
        files.sort()
        if not files:
            self.error.emit("이미지가 없습니다.")
            return
        os.makedirs(self.out_dir, exist_ok=True)
        for i, fp in enumerate(files):
            frame = cv2.imread(fp)
            if frame is None:
                continue
            h, w = frame.shape[:2]
            res = self._infer(frame)
            txt_path = os.path.join(self.out_dir, os.path.splitext(os.path.basename(fp))[0] + ".txt")
            self._write_label(txt_path, res, h, w)
            self.progress.emit(i + 1, len(files))
        self.finished_ok.emit(self.out_dir)

    def _run_video(self):
        vid_exts = (".mp4", ".avi", ".mkv", ".mov", ".wmv")
        # src_path가 폴더면 내부 비디오 수집, 파일이면 단일 처리
        if os.path.isdir(self.src_path):
            videos = sorted(
                f for f in glob.glob(os.path.join(self.src_path, "*"))
                if os.path.splitext(f)[1].lower() in vid_exts
            )
        else:
            videos = [self.src_path]
        if not videos:
            self.error.emit("비디오 파일이 없습니다.")
            return

        # 전체 프레임 수 사전 계산 (진행률용)
        total_frames = 0
        for vp in videos:
            cap = cv2.VideoCapture(vp)
            total_frames += int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) // self.frame_interval
            cap.release()
        if total_frames <= 0:
            total_frames = 1

        os.makedirs(self.out_dir, exist_ok=True)
        img_dir = os.path.join(self.out_dir, "images") if self.save_frames else None
        lbl_dir = os.path.join(self.out_dir, "labels")
        os.makedirs(lbl_dir, exist_ok=True)
        if img_dir:
            os.makedirs(img_dir, exist_ok=True)

        processed = 0
        for vp in videos:
            cap = cv2.VideoCapture(vp)
            if not cap.isOpened():
                continue
            vid_name = os.path.splitext(os.path.basename(vp))[0]
            idx = 0
            while True:
                ret, frame = cap.read()
                if not ret:
                    break
                if idx % self.frame_interval == 0:
                    h, w = frame.shape[:2]
                    res = self._infer(frame)
                    fname = f"{vid_name}_{idx:06d}"
                    self._write_label(os.path.join(lbl_dir, fname + ".txt"), res, h, w)
                    if img_dir:
                        cv2.imwrite(os.path.join(img_dir, fname + ".jpg"), frame)
                    processed += 1
                    self.progress.emit(processed, total_frames)
                idx += 1
            cap.release()
        self.finished_ok.emit(self.out_dir)

    def run(self):
        try:
            if self.source_type == "video":
                self._run_video()
            else:
                self._run_images()
        except Exception as e:
            self.error.emit(str(e))


_IGNORE_CLASSES = {99}


# ------------------------------------------------------------------ #
# 멀티모델 평가 워커
# ------------------------------------------------------------------ #
class _EvalWorker(QThread):
    """여러 모델을 순차 추론 → GT 대비 평가"""
    progress = Signal(str, int, int)        # model_name, current, total
    model_done = Signal(str, dict)          # model_name, metrics dict
    all_done = Signal()
    error = Signal(str)

    def __init__(self, model_entries, img_dir, gt_dir, conf, classmap=None,
                 per_model_mappings=None, mapped_only=True):
        super().__init__()
        self.model_entries = model_entries
        self.img_dir = img_dir
        self.gt_dir = gt_dir
        self.conf = conf
        self.classmap = classmap or {}
        self.per_model_mappings = per_model_mappings or {}
        self.mapped_only = mapped_only
        self._stop = False

    def stop(self):
        self._stop = True

    def run(self):
        try:
            exts = ("*.jpg", "*.jpeg", "*.png", "*.bmp")
            files = []
            for e in exts:
                files.extend(glob.glob(os.path.join(self.img_dir, e)))
            files.sort()
            if not files:
                self.error.emit("이미지가 없습니다.")
                return

            gt_data = _load_yolo_labels(self.gt_dir, self.img_dir)
            gt_raw = _load_yolo_labels(self.gt_dir, self.img_dir, ignore_classes=set())

            for model_path, model_type in self.model_entries:
                name = os.path.basename(model_path)
                try:
                    mi = load_model(model_path, model_type=model_type)
                except Exception as exc:
                    self.error.emit(f"{name} 로드 실패: {exc}")
                    continue

                # 추론 → pred_data 생성
                model_stem = os.path.splitext(name)[0]
                vis_dir = os.path.join(os.path.dirname(self.img_dir), "eval_results", model_stem)
                os.makedirs(vis_dir, exist_ok=True)
                # 모델 지원 클래스 + GT 리매핑 (per-model mapping 우선)
                mapping = self.per_model_mappings.get(name, {})
                if mapping:
                    # 매핑 다이얼로그 결과 사용
                    allowed_gt = set(mapping.values()) if self.mapped_only else None
                    gt_mapped = gt_data
                    if allowed_gt:
                        gt_mapped = {s: [b for b in boxes if b[0] in allowed_gt]
                                     for s, boxes in gt_data.items()}
                else:
                    # fallback: 기존 방식
                    model_cmap = _build_model_classmap(mi.names) if mi.names else {}
                    if model_type == "darknet":
                        allowed = set(UNIFIED_NAMES.keys())
                    else:
                        allowed = set(model_cmap.values()) if model_cmap else None
                    gt_mapped = _remap_gt(gt_data, self.classmap, allowed) if allowed else gt_data
                pred_data = {}
                for i, fp in enumerate(files):
                    if self._stop:
                        self.all_done.emit()
                        return
                    frame = cv2.imread(fp)
                    if frame is None:
                        continue
                    stem = os.path.splitext(os.path.basename(fp))[0]
                    # class 99 마스킹
                    if stem in gt_raw and any(b[0] in _IGNORE_CLASSES for b in gt_raw[stem]):
                        frame = _mask_class99(frame, gt_raw[stem])
                    h, w = frame.shape[:2]
                    res = run_inference(mi, frame, self.conf)
                    if model_type == "darknet":
                        res = convert_darknet_to_unified(res)
                    stem = os.path.splitext(os.path.basename(fp))[0]
                    boxes = []
                    for box, score, cid in zip(res.boxes, res.scores, res.class_ids):
                        cid = int(cid)
                        # per-model mapping으로 클래스 리매핑
                        if mapping:
                            if cid in mapping:
                                cid = mapping[cid]
                            elif self.mapped_only:
                                continue
                        elif model_type != "darknet" and model_cmap:
                            cid = model_cmap.get(cid, -1)
                            if cid < 0:
                                continue
                        if cid in _IGNORE_CLASSES:
                            continue
                        x1, y1, x2, y2 = box
                        cx = ((x1 + x2) / 2) / w
                        cy = ((y1 + y2) / 2) / h
                        bw = (x2 - x1) / w
                        bh = (y2 - y1) / h
                        boxes.append((cid, cx, cy, bw, bh, float(score)))
                    pred_data[stem] = boxes

                    # GT vs Pred: 클래스 불일치만 시각화
                    vis = frame.copy()
                    gt_boxes_img = gt_mapped.get(stem, [])
                    gt_matched = [False] * len(gt_boxes_img)
                    pred_matched = [False] * len(boxes)
                    # 모든 쌍 IoU 계산 → 같은 클래스 우선 매칭
                    pairs = []
                    for pi, (pc, pcx, pcy, pbw, pbh, _ps) in enumerate(boxes):
                        pbox = _yolo_to_xyxy(pcx, pcy, pbw, pbh)
                        for gi, (gc, gcx, gcy, gbw, gbh) in enumerate(gt_boxes_img):
                            gbox = _yolo_to_xyxy(gcx, gcy, gbw, gbh)
                            iou = _compute_iou(pbox, gbox)
                            if iou > 0.1:
                                pairs.append((iou, pi, gi, pc == gc))
                    # pass 1: 같은 클래스끼리 매칭
                    for iou, pi, gi, same in sorted(pairs, key=lambda x: -x[0]):
                        if not same or pred_matched[pi] or gt_matched[gi]:
                            continue
                        pred_matched[pi] = True
                        gt_matched[gi] = True
                    # pass 2: 남은 것 크로스 매칭
                    for iou, pi, gi, same in sorted(pairs, key=lambda x: -x[0]):
                        if pred_matched[pi] or gt_matched[gi]:
                            continue
                        pred_matched[pi] = True
                        gt_matched[gi] = True
                        pc = boxes[pi][0]
                        gc = gt_boxes_img[gi][0]
                        if pc != gc:
                            # 클래스 불일치 — 둘 다 그리기
                            gc, gcx, gcy, gbw, gbh = gt_boxes_img[gi]
                            x1i = int((gcx - gbw/2)*w); y1i = int((gcy - gbh/2)*h)
                            x2i = int((gcx + gbw/2)*w); y2i = int((gcy + gbh/2)*h)
                            cv2.rectangle(vis, (x1i,y1i), (x2i,y2i), (0,200,0), 3)
                            cv2.putText(vis, f"GT:{UNIFIED_NAMES.get(gc,str(gc))}", (x1i, max(y1i-6,14)),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,200,0), 2)
                            pc, pcx, pcy, pbw, pbh, _ps = boxes[pi]
                            x1i = int((pcx - pbw/2)*w); y1i = int((pcy - pbh/2)*h)
                            x2i = int((pcx + pbw/2)*w); y2i = int((pcy + pbh/2)*h)
                            cv2.rectangle(vis, (x1i,y1i), (x2i,y2i), (0,0,220), 3)
                            cv2.putText(vis, f"Pred:{UNIFIED_NAMES.get(pc,str(pc))}", (x1i, max(y1i-6,14)),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,0,220), 2)
                    # 미매칭 GT (FN)
                    for gi, matched in enumerate(gt_matched):
                        if not matched:
                            gc, gcx, gcy, gbw, gbh = gt_boxes_img[gi]
                            x1i = int((gcx - gbw/2)*w); y1i = int((gcy - gbh/2)*h)
                            x2i = int((gcx + gbw/2)*w); y2i = int((gcy + gbh/2)*h)
                            cv2.rectangle(vis, (x1i,y1i), (x2i,y2i), (0,200,0), 3)
                            cv2.putText(vis, f"FN:{UNIFIED_NAMES.get(gc,str(gc))}", (x1i, max(y1i-6,14)),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,200,0), 2)
                    # 미매칭 Pred (FP)
                    for pi, matched in enumerate(pred_matched):
                        if not matched:
                            pc, pcx, pcy, pbw, pbh, _ps = boxes[pi]
                            x1i = int((pcx - pbw/2)*w); y1i = int((pcy - pbh/2)*h)
                            x2i = int((pcx + pbw/2)*w); y2i = int((pcy + pbh/2)*h)
                            cv2.rectangle(vis, (x1i,y1i), (x2i,y2i), (0,0,220), 3)
                            cv2.putText(vis, f"FP:{UNIFIED_NAMES.get(pc,str(pc))}", (x1i, max(y1i-6,14)),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,0,220), 2)
                    cv2.imwrite(os.path.join(vis_dir, stem + ".jpg"), vis)

                    self.progress.emit(name, i + 1, len(files))

                # 모델 지원 클래스 필터링 후 평가
                if mapping and self.mapped_only:
                    allowed_gt = set(mapping.values())
                    pred_eval = {s: [b for b in boxes if b[0] in allowed_gt]
                                 for s, boxes in pred_data.items()}
                elif not mapping:
                    allowed = set(model_cmap.values()) if (not mapping and model_cmap) else None
                    pred_eval = _filter_pred(pred_data, allowed) if allowed else pred_data
                else:
                    pred_eval = pred_data

                res50 = evaluate_dataset(gt_mapped, pred_eval, 0.5)
                map5095 = evaluate_map50_95(gt_mapped, pred_eval)
                ov = res50.get("__overall__", {})
                self.model_done.emit(name, {
                    "mAP50": ov.get("ap", 0),
                    "mAP50_95": map5095,
                    "precision": ov.get("precision", 0),
                    "recall": ov.get("recall", 0),
                    "f1": ov.get("f1", 0),
                    "detail": res50,
                })

            self.all_done.emit()
        except Exception as e:
            self.error.emit(str(e))


# ------------------------------------------------------------------ #
# 평가 메트릭 계산
# ------------------------------------------------------------------ #
def _compute_iou(box1, box2):
    x1 = max(box1[0], box2[0])
    y1 = max(box1[1], box2[1])
    x2 = min(box1[2], box2[2])
    y2 = min(box1[3], box2[3])
    inter = max(0, x2 - x1) * max(0, y2 - y1)
    a1 = (box1[2] - box1[0]) * (box1[3] - box1[1])
    a2 = (box2[2] - box2[0]) * (box2[3] - box2[1])
    return inter / (a1 + a2 - inter + 1e-9)


def _load_yolo_labels(label_dir, img_dir, ignore_classes=None):
    """YOLO txt 라벨 로드 → {filename: [(cid, cx, cy, w, h), ...]}"""
    if ignore_classes is None:
        ignore_classes = _IGNORE_CLASSES
    data = {}
    exts = ("*.jpg", "*.jpeg", "*.png", "*.bmp")
    imgs = []
    for e in exts:
        imgs.extend(glob.glob(os.path.join(img_dir, e)))
    for img_path in imgs:
        stem = os.path.splitext(os.path.basename(img_path))[0]
        txt = os.path.join(label_dir, stem + ".txt")
        boxes = []
        if os.path.isfile(txt):
            with open(txt) as f:
                for line in f:
                    parts = line.strip().split()
                    if len(parts) >= 5:
                        cid = int(parts[0])
                        if cid in ignore_classes:
                            continue
                        cx, cy, w, h = float(parts[1]), float(parts[2]), float(parts[3]), float(parts[4])
                        boxes.append((cid, cx, cy, w, h))
        data[stem] = boxes
    return data


def _mask_class99(frame, gt_all):
    """99번 영역 검정 마스킹 후, 다른 클래스 영역 원본 복원"""
    h, w = frame.shape[:2]
    masked = frame.copy()
    # 1) class 99 영역 검정
    for cid, cx, cy, bw, bh in gt_all:
        if cid not in _IGNORE_CLASSES:
            continue
        x1 = max(0, int((cx - bw / 2) * w))
        y1 = max(0, int((cy - bh / 2) * h))
        x2 = min(w, int((cx + bw / 2) * w))
        y2 = min(h, int((cy + bh / 2) * h))
        masked[y1:y2, x1:x2] = 0
    # 2) 다른 클래스 영역 원본 복원
    for cid, cx, cy, bw, bh in gt_all:
        if cid in _IGNORE_CLASSES:
            continue
        x1 = max(0, int((cx - bw / 2) * w))
        y1 = max(0, int((cy - bh / 2) * h))
        x2 = min(w, int((cx + bw / 2) * w))
        y2 = min(h, int((cy + bh / 2) * h))
        masked[y1:y2, x1:x2] = frame[y1:y2, x1:x2]
    return masked


def _yolo_to_xyxy(cx, cy, w, h):
    return (cx - w / 2, cy - h / 2, cx + w / 2, cy + h / 2)


def _parse_classmap(text):
    """'id: name' 텍스트 → {gt_id: unified_id} 매핑 생성"""
    name_to_unified = {v.lower(): k for k, v in UNIFIED_NAMES.items()}
    mapping = {}
    for line in text.strip().splitlines():
        line = line.strip()
        if ":" not in line:
            continue
        gid_str, name = line.split(":", 1)
        try:
            gid = int(gid_str.strip())
        except ValueError:
            continue
        name = name.strip().lower()
        if name in name_to_unified:
            mapping[gid] = name_to_unified[name]
    return mapping


def _build_model_classmap(model_names):
    """model_info.names → {model_id: unified_id} 이름 기반 매핑"""
    name_to_unified = {v.lower(): k for k, v in UNIFIED_NAMES.items()}
    # nomask → mask로 통합, front/back/side 무시
    name_to_unified["nomask"] = 9   # mask
    _SKIP_NAMES = {"front", "back", "side"}
    mapping = {}
    for mid, mname in model_names.items():
        n = mname.strip().lower()
        if n in _SKIP_NAMES:
            continue
        uid = name_to_unified.get(n)
        if uid is not None:
            mapping[mid] = uid
    return mapping


def _remap_gt(gt_data, classmap, allowed_classes):
    """GT 클래스 ID를 unified ID로 리매핑 + 허용 클래스만 필터"""
    remapped = {}
    for stem, boxes in gt_data.items():
        new_boxes = []
        for b in boxes:
            uid = classmap.get(b[0], b[0])
            if uid in allowed_classes:
                new_boxes.append((uid, *b[1:]))
        remapped[stem] = new_boxes
    return remapped


def _filter_pred(pred_data, allowed_classes):
    """pred에서 허용 클래스만 필터"""
    filtered = {}
    for stem, boxes in pred_data.items():
        filtered[stem] = [b for b in boxes if b[0] in allowed_classes]
    return filtered


def _compute_ap(recalls, precisions):
    """101-point interpolation AP"""
    mrec = np.concatenate(([0.0], recalls, [1.0]))
    mpre = np.concatenate(([1.0], precisions, [0.0]))
    for i in range(len(mpre) - 2, -1, -1):
        mpre[i] = max(mpre[i], mpre[i + 1])
    ap = 0.0
    for t in np.linspace(0, 1, 101):
        p = mpre[mrec >= t]
        ap += (p.max() if len(p) > 0 else 0.0)
    return ap / 101.0


def evaluate_dataset(gt_data, pred_data, iou_thres=0.5):
    """단일 IoU threshold에서 per-class AP, Precision, Recall 계산"""
    all_classes = set()
    for boxes in list(gt_data.values()) + list(pred_data.values()):
        for b in boxes:
            all_classes.add(b[0])
    if not all_classes:
        return {}

    results = {}
    total_tp = total_fp = total_fn = 0

    for cid in sorted(all_classes):
        # 모든 이미지에서 해당 클래스의 pred/gt 수집
        preds = []
        gt_per_img = {}
        for stem in set(list(gt_data.keys()) + list(pred_data.keys())):
            gt_boxes = [_yolo_to_xyxy(b[1], b[2], b[3], b[4]) for b in gt_data.get(stem, []) if b[0] == cid]
            gt_per_img[stem] = {"boxes": gt_boxes, "matched": [False] * len(gt_boxes)}
            for b in pred_data.get(stem, []):
                if b[0] == cid:
                    score = b[5] if len(b) > 5 else 0.0
                    preds.append((score, stem, _yolo_to_xyxy(b[1], b[2], b[3], b[4])))
        preds.sort(key=lambda x: -x[0])  # confidence 내림차순

        n_gt = sum(len(v["boxes"]) for v in gt_per_img.values())
        if n_gt == 0 and len(preds) == 0:
            continue

        tp_list = []
        for _score, stem, pbox in preds:
            best_iou = 0
            best_j = -1
            for j, gbox in enumerate(gt_per_img[stem]["boxes"]):
                iou = _compute_iou(pbox, gbox)
                if iou > best_iou:
                    best_iou = iou
                    best_j = j
            if best_iou >= iou_thres and not gt_per_img[stem]["matched"][best_j]:
                gt_per_img[stem]["matched"][best_j] = True
                tp_list.append(1)
            else:
                tp_list.append(0)

        tp_arr = np.array(tp_list)
        tp_cum = np.cumsum(tp_arr)
        fp_cum = np.cumsum(1 - tp_arr)
        recalls = tp_cum / max(n_gt, 1)
        precisions = tp_cum / (tp_cum + fp_cum + 1e-9)

        ap = _compute_ap(recalls, precisions) if len(recalls) > 0 else 0.0
        tp_total = int(tp_arr.sum())
        fp_total = len(preds) - tp_total
        fn_total = n_gt - tp_total

        total_tp += tp_total
        total_fp += fp_total
        total_fn += fn_total

        prec = tp_total / (tp_total + fp_total + 1e-9)
        rec = tp_total / (n_gt + 1e-9)
        f1 = 2 * prec * rec / (prec + rec + 1e-9)
        results[cid] = {"ap": ap, "precision": prec, "recall": rec, "f1": f1, "tp": tp_total, "fp": fp_total, "fn": fn_total}

    # overall
    prec_all = total_tp / (total_tp + total_fp + 1e-9)
    rec_all = total_tp / (total_tp + total_fn + 1e-9)
    f1_all = 2 * prec_all * rec_all / (prec_all + rec_all + 1e-9)
    acc_all = total_tp / (total_tp + total_fp + total_fn + 1e-9)
    mean_ap = np.mean([v["ap"] for v in results.values()]) if results else 0.0
    results["__overall__"] = {"ap": mean_ap, "precision": prec_all, "recall": rec_all, "f1": f1_all, "accuracy": acc_all}
    return results


def evaluate_map50_95(gt_data, pred_data):
    """mAP50:95 계산 (0.50~0.95, step 0.05)"""
    aps = []
    for iou_t in np.arange(0.5, 1.0, 0.05):
        res = evaluate_dataset(gt_data, pred_data, iou_t)
        if "__overall__" in res:
            aps.append(res["__overall__"]["ap"])
        else:
            aps.append(0.0)
    return float(np.mean(aps)) if aps else 0.0


# ------------------------------------------------------------------ #
# Classification 평가
# ------------------------------------------------------------------ #

def _load_cls_labels(label_dir, img_dir):
    """Classification 라벨 로드 → {stem: class_id}
    지원 포맷: (1) 이미지별 txt (class_id 한 줄), (2) 폴더명이 클래스명"""
    data = {}
    exts = ("*.jpg", "*.jpeg", "*.png", "*.bmp")
    imgs = []
    for e in exts:
        imgs.extend(glob.glob(os.path.join(img_dir, e)))
    for img_path in imgs:
        stem = os.path.splitext(os.path.basename(img_path))[0]
        txt = os.path.join(label_dir, stem + ".txt")
        if os.path.isfile(txt):
            with open(txt) as f:
                line = f.readline().strip()
                if line:
                    try:
                        data[stem] = int(line.split()[0])
                    except ValueError:
                        pass
    # 폴더 구조 방식: img_dir 자체가 클래스 폴더 안에 있는 경우
    if not data:
        parent = os.path.basename(os.path.dirname(img_dir))
        # 상위 폴더의 서브폴더들을 클래스로 사용
        parent_dir = os.path.dirname(img_dir)
        subdirs = sorted([d for d in os.listdir(parent_dir)
                         if os.path.isdir(os.path.join(parent_dir, d))])
        if subdirs:
            cls_map = {name: i for i, name in enumerate(subdirs)}
            folder_name = os.path.basename(img_dir)
            if folder_name in cls_map:
                cid = cls_map[folder_name]
                for img_path in imgs:
                    stem = os.path.splitext(os.path.basename(img_path))[0]
                    data[stem] = cid
    return data


def evaluate_classification(gt_data, pred_data, num_classes):
    """Classification 평가: Accuracy, per-class Precision/Recall/F1
    gt_data: {stem: class_id}, pred_data: {stem: (class_id, confidence)}"""
    if not gt_data or not pred_data:
        return {}
    y_true, y_pred = [], []
    for stem in gt_data:
        if stem in pred_data:
            y_true.append(gt_data[stem])
            y_pred.append(pred_data[stem][0])
    if not y_true:
        return {}
    y_true, y_pred = np.array(y_true), np.array(y_pred)
    accuracy = float((y_true == y_pred).mean())

    results = {}
    all_classes = sorted(set(y_true.tolist() + y_pred.tolist()))
    for cid in all_classes:
        tp = int(((y_true == cid) & (y_pred == cid)).sum())
        fp = int(((y_true != cid) & (y_pred == cid)).sum())
        fn = int(((y_true == cid) & (y_pred != cid)).sum())
        prec = tp / (tp + fp + 1e-9)
        rec = tp / (tp + fn + 1e-9)
        f1 = 2 * prec * rec / (prec + rec + 1e-9)
        results[cid] = {"precision": prec, "recall": rec, "f1": f1, "tp": tp, "fp": fp, "fn": fn}

    macro_prec = np.mean([v["precision"] for v in results.values()])
    macro_rec = np.mean([v["recall"] for v in results.values()])
    macro_f1 = np.mean([v["f1"] for v in results.values()])
    results["__overall__"] = {
        "accuracy": accuracy, "precision": float(macro_prec),
        "recall": float(macro_rec), "f1": float(macro_f1),
    }
    return results


class _ClsEvalWorker(QThread):
    """Classification 모델 평가 워커"""
    progress = Signal(str, int, int)
    model_done = Signal(str, dict)
    all_done = Signal()
    error = Signal(str)

    def __init__(self, model_entries, img_dir, gt_dir, top_k=5):
        super().__init__()
        self.model_entries = model_entries
        self.img_dir = img_dir
        self.gt_dir = gt_dir
        self.top_k = top_k
        self._stop = False

    def stop(self):
        self._stop = True

    def run(self):
        try:
            exts = ("*.jpg", "*.jpeg", "*.png", "*.bmp")
            files = []
            for e in exts:
                files.extend(glob.glob(os.path.join(self.img_dir, e)))
            files.sort()
            if not files:
                self.error.emit("이미지가 없습니다.")
                return

            gt_data = _load_cls_labels(self.gt_dir, self.img_dir)

            for model_path, model_type in self.model_entries:
                name = os.path.basename(model_path)
                try:
                    mi = load_model(model_path, model_type=model_type)
                except Exception as exc:
                    self.error.emit(f"{name} 로드 실패: {exc}")
                    continue

                pred_data = {}
                for i, fp in enumerate(files):
                    if self._stop:
                        self.all_done.emit()
                        return
                    frame = cv2.imread(fp)
                    if frame is None:
                        continue
                    stem = os.path.splitext(os.path.basename(fp))[0]
                    res = run_classification(mi, frame, self.top_k)
                    pred_data[stem] = (res.class_id, res.confidence)
                    self.progress.emit(name, i + 1, len(files))

                num_classes = len(mi.names) if mi.names else 0
                metrics = evaluate_classification(gt_data, pred_data, num_classes)
                ov = metrics.get("__overall__", {})
                self.model_done.emit(name, {
                    "accuracy": ov.get("accuracy", 0),
                    "precision": ov.get("precision", 0),
                    "recall": ov.get("recall", 0),
                    "f1": ov.get("f1", 0),
                    "detail": metrics,
                })

            self.all_done.emit()
        except Exception as e:
            self.error.emit(str(e))


# ------------------------------------------------------------------ #
# 평가 탭 위젯
# ------------------------------------------------------------------ #
class EvaluationTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._model_info: ModelInfo | None = None
        self._worker = None
        self._last_mappings = {}      # 이전 매핑 기억
        self._last_mapped_only = True
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # === Pseudo Labeling ===
        grp1 = QGroupBox("Pseudo Labeling")
        g1 = QVBoxLayout(grp1)

        row_model = QHBoxLayout()
        row_model.addWidget(QLabel("모델:"))
        self._le_model = QLineEdit()
        self._le_model.setReadOnly(True)
        row_model.addWidget(self._le_model, 1)
        btn_model = QPushButton("찾아보기")
        btn_model.clicked.connect(self._browse_model)
        row_model.addWidget(btn_model)
        self._combo_type = QComboBox()
        from core.model_loader import MODEL_TYPES
        from core.app_config import AppConfig as _AC
        for key, label in MODEL_TYPES.items():
            self._combo_type.addItem(label, key)
        for name in _AC().custom_model_types:
            self._combo_type.addItem(name, f"custom:{name}")
        row_model.addWidget(self._combo_type)
        g1.addLayout(row_model)

        row_img = QHBoxLayout()
        row_img.addWidget(QLabel("소스:"))
        self._combo_source = QComboBox()
        self._combo_source.addItems(["이미지 폴더", "비디오"])
        self._combo_source.currentIndexChanged.connect(self._on_source_changed)
        row_img.addWidget(self._combo_source)
        self._le_img_dir = QLineEdit()
        row_img.addWidget(self._le_img_dir, 1)
        self._btn_browse_src = QPushButton("찾아보기")
        self._btn_browse_src.clicked.connect(self._browse_pseudo_src)
        row_img.addWidget(self._btn_browse_src)
        g1.addLayout(row_img)

        # 비디오 옵션 행
        self._row_vid_opts = QHBoxLayout()
        self._row_vid_opts.addWidget(QLabel("프레임 간격:"))
        self._spin_interval = QDoubleSpinBox()
        self._spin_interval.setDecimals(0)
        self._spin_interval.setRange(1, 9999)
        self._spin_interval.setValue(30)
        self._spin_interval.setToolTip("N프레임마다 1장 추출 (예: 30 = 1초당 1장 @30fps)")
        self._row_vid_opts.addWidget(self._spin_interval)
        self._chk_save_frames = QCheckBox("프레임 이미지 저장")
        self._chk_save_frames.setChecked(True)
        self._row_vid_opts.addWidget(self._chk_save_frames)
        self._row_vid_opts.addStretch()
        g1.addLayout(self._row_vid_opts)
        # 초기 상태: 비디오 옵션 숨김
        self._set_vid_opts_visible(False)

        row_out = QHBoxLayout()
        row_out.addWidget(QLabel("출력 폴더:"))
        self._le_out_dir = QLineEdit()
        row_out.addWidget(self._le_out_dir, 1)
        btn_out = QPushButton("찾아보기")
        btn_out.clicked.connect(lambda: self._browse_dir(self._le_out_dir))
        row_out.addWidget(btn_out)
        g1.addLayout(row_out)

        row_conf = QHBoxLayout()
        row_conf.addWidget(QLabel("Confidence:"))
        self._spin_conf = QDoubleSpinBox()
        self._spin_conf.setRange(0.01, 1.0)
        self._spin_conf.setValue(0.25)
        self._spin_conf.setSingleStep(0.05)
        row_conf.addWidget(self._spin_conf)
        row_conf.addStretch()
        self._btn_pseudo = QPushButton("Pseudo Labeling 실행")
        self._btn_pseudo.clicked.connect(self._run_pseudo)
        row_conf.addWidget(self._btn_pseudo)
        g1.addLayout(row_conf)

        self._prog = QProgressBar()
        g1.addWidget(self._prog)
        layout.addWidget(grp1)

        # === 정량적 평가 ===
        grp2 = QGroupBox("정량적 평가")
        g2 = QVBoxLayout(grp2)

        # 태스크 선택
        row_task = QHBoxLayout()
        row_task.addWidget(QLabel("태스크:"))
        self._combo_task = QComboBox()
        self._combo_task.addItems(["Detection", "Classification"])
        self._combo_task.currentTextChanged.connect(self._on_task_changed)
        row_task.addWidget(self._combo_task)
        row_task.addStretch()
        g2.addLayout(row_task)

        row_gt = QHBoxLayout()
        row_gt.addWidget(QLabel("GT 라벨 폴더:"))
        self._le_gt = QLineEdit()
        row_gt.addWidget(self._le_gt, 1)
        btn_gt = QPushButton("찾아보기")
        btn_gt.clicked.connect(lambda: self._browse_dir(self._le_gt))
        row_gt.addWidget(btn_gt)
        g2.addLayout(row_gt)

        row_eval_img = QHBoxLayout()
        row_eval_img.addWidget(QLabel("이미지 폴더:"))
        self._le_eval_img = QLineEdit()
        row_eval_img.addWidget(self._le_eval_img, 1)
        btn_eval_img = QPushButton("찾아보기")
        btn_eval_img.clicked.connect(lambda: self._browse_dir(self._le_eval_img))
        row_eval_img.addWidget(btn_eval_img)
        g2.addLayout(row_eval_img)

        # GT 클래스 매핑
        row_cls = QHBoxLayout()
        row_cls.addWidget(QLabel("GT 클래스 매핑:"))
        self._te_classmap = QTextEdit()
        self._te_classmap.setMaximumHeight(100)
        self._te_classmap.setPlaceholderText("0: person\n1: face\n2: red-sign\n3: wheelchair\n4: cane")
        # 기본값: unified names
        default_map = "\n".join(f"{k}: {v}" for k, v in sorted(UNIFIED_NAMES.items()))
        self._te_classmap.setPlainText(default_map)
        row_cls.addWidget(self._te_classmap, 1)
        g2.addLayout(row_cls)

        # 모델 리스트
        row_emodels = QHBoxLayout()
        row_emodels.addWidget(QLabel("모델:"))
        self._eval_model_list = QTableWidget()
        self._eval_model_list.setColumnCount(2)
        self._eval_model_list.setHorizontalHeaderLabels(["모델 경로", "타입"])
        self._eval_model_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self._eval_model_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self._eval_model_list.setMaximumHeight(120)
        row_emodels.addWidget(self._eval_model_list, 1)
        btn_col = QVBoxLayout()
        btn_add = QPushButton("+")
        btn_add.setFixedWidth(30)
        btn_add.clicked.connect(self._add_eval_model)
        btn_rm = QPushButton("−")
        btn_rm.setFixedWidth(30)
        btn_rm.clicked.connect(self._remove_eval_model)
        btn_col.addWidget(btn_add)
        btn_col.addWidget(btn_rm)
        btn_col.addStretch()
        row_emodels.addLayout(btn_col)
        g2.addLayout(row_emodels)

        row_eval_run = QHBoxLayout()
        row_eval_run.addWidget(QLabel("Confidence:"))
        self._spin_eval_conf = QDoubleSpinBox()
        self._spin_eval_conf.setRange(0.01, 1.0)
        self._spin_eval_conf.setValue(0.25)
        self._spin_eval_conf.setSingleStep(0.05)
        row_eval_run.addWidget(self._spin_eval_conf)
        row_eval_run.addStretch()
        self._btn_eval = QPushButton("평가 실행")
        self._btn_eval.clicked.connect(self._run_eval)
        row_eval_run.addWidget(self._btn_eval)
        self._btn_eval_stop = QPushButton("중지")
        self._btn_eval_stop.setEnabled(False)
        self._btn_eval_stop.clicked.connect(self._stop_eval)
        row_eval_run.addWidget(self._btn_eval_stop)
        self._btn_export = QPushButton("Excel 내보내기")
        self._btn_export.clicked.connect(self._export_excel)
        row_eval_run.addWidget(self._btn_export)
        self._btn_export_html = QPushButton("HTML 리포트")
        self._btn_export_html.clicked.connect(self._export_html)
        row_eval_run.addWidget(self._btn_export_html)
        g2.addLayout(row_eval_run)

        self._eval_prog = QProgressBar()
        g2.addWidget(self._eval_prog)

        # 결과 테이블 (모델별 요약)
        self._table = QTableWidget()
        self._table.setColumnCount(7)
        self._table.setHorizontalHeaderLabels(["모델", "mAP@50", "mAP@50:95", "Precision", "Recall", "F1", ""])
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        g2.addWidget(self._table)

        # 요약
        self._lbl_summary = QTextEdit()
        self._lbl_summary.setReadOnly(True)
        self._lbl_summary.setMaximumHeight(80)
        g2.addWidget(self._lbl_summary)

        layout.addWidget(grp2)

    # ---- helpers ----
    def _browse_model(self):
        path, _ = QFileDialog.getOpenFileName(self, "모델 선택", "Models", "ONNX (*.onnx)")
        if path:
            self._le_model.setText(path)
            model_type = self._combo_type.currentData() or "yolo"
            try:
                self._model_info = load_model(path, model_type=model_type)
            except Exception as e:
                QMessageBox.critical(self, "모델 로드 실패", str(e))

    def _browse_dir(self, le: QLineEdit):
        d = QFileDialog.getExistingDirectory(self, "폴더 선택")
        if d:
            le.setText(d)

    def _set_vid_opts_visible(self, visible):
        for i in range(self._row_vid_opts.count()):
            w = self._row_vid_opts.itemAt(i).widget()
            if w:
                w.setVisible(visible)

    def _on_source_changed(self, idx):
        is_video = idx == 1
        self._set_vid_opts_visible(is_video)

    def _browse_pseudo_src(self):
        if self._combo_source.currentIndex() == 0:
            self._browse_dir(self._le_img_dir)
        else:
            # 비디오: 파일 또는 폴더 선택
            from PySide6.QtWidgets import QMenu
            menu = QMenu(self)
            act_file = menu.addAction("비디오 파일 선택")
            act_dir = menu.addAction("비디오 폴더 선택")
            action = menu.exec(self._btn_browse_src.mapToGlobal(self._btn_browse_src.rect().bottomLeft()))
            if action == act_file:
                path, _ = QFileDialog.getOpenFileName(
                    self, "비디오 선택", "",
                    "Video (*.mp4 *.avi *.mkv *.mov *.wmv)")
                if path:
                    self._le_img_dir.setText(path)
            elif action == act_dir:
                self._browse_dir(self._le_img_dir)

    # ---- pseudo labeling ----
    def _run_pseudo(self):
        if not self._model_info or not self._model_info.session:
            QMessageBox.warning(self, "알림", "모델을 먼저 선택하세요.")
            return
        src_path = self._le_img_dir.text()
        if not src_path or not os.path.exists(src_path):
            QMessageBox.warning(self, "알림", "소스 경로를 선택하세요.")
            return
        out_dir = self._le_out_dir.text()
        if not out_dir:
            parent = os.path.dirname(src_path) if os.path.isfile(src_path) else os.path.dirname(src_path)
            out_dir = os.path.join(parent, "pseudo_labels")
            self._le_out_dir.setText(out_dir)

        is_video = self._combo_source.currentIndex() == 1
        self._btn_pseudo.setEnabled(False)
        self._worker = _PseudoWorker(
            self._model_info, src_path, out_dir, self._spin_conf.value(),
            source_type="video" if is_video else "image",
            frame_interval=int(self._spin_interval.value()) if is_video else 1,
            save_frames=self._chk_save_frames.isChecked() if is_video else False,
        )
        self._worker.progress.connect(lambda c, t: self._prog.setValue(int(c / t * 100) if t else 0))
        self._worker.finished_ok.connect(self._on_pseudo_done)
        self._worker.error.connect(lambda e: (QMessageBox.critical(self, "오류", e), self._btn_pseudo.setEnabled(True)))
        self._prog.setValue(0)
        self._worker.start()

    def _on_pseudo_done(self, out_dir):
        self._btn_pseudo.setEnabled(True)
        self._prog.setValue(100)
        QMessageBox.information(self, "완료", f"Pseudo labeling 완료\n{out_dir}")

    # ---- 평가 모델 관리 ----
    def _add_eval_model(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "모델 선택", "Models", "ONNX (*.onnx)")
        for p in paths:
            row = self._eval_model_list.rowCount()
            self._eval_model_list.insertRow(row)
            self._eval_model_list.setItem(row, 0, QTableWidgetItem(p))
            combo = QComboBox()
            from core.model_loader import MODEL_TYPES
            from core.app_config import AppConfig as _AC
            for key, label in MODEL_TYPES.items():
                combo.addItem(label, key)
            for name in _AC().custom_model_types:
                combo.addItem(name, f"custom:{name}")
            self._eval_model_list.setCellWidget(row, 1, combo)

    def _remove_eval_model(self):
        rows = sorted(set(i.row() for i in self._eval_model_list.selectedIndexes()), reverse=True)
        for r in rows:
            self._eval_model_list.removeRow(r)

    # ---- 평가 ----
    def _on_task_changed(self, task):
        """태스크 변경 시 UI 업데이트"""
        is_det = (task == "Detection")
        self._te_classmap.setVisible(is_det)
        self._spin_eval_conf.setVisible(is_det)
        # 라벨 텍스트 변경
        if is_det:
            self._table.setColumnCount(7)
            self._table.setHorizontalHeaderLabels(["모델", "mAP@50", "mAP@50:95", "Precision", "Recall", "F1", ""])
        else:
            self._table.setColumnCount(6)
            self._table.setHorizontalHeaderLabels(["모델", "Accuracy", "Precision", "Recall", "F1", ""])
        self._table.setRowCount(0)

    def _run_eval(self):
        gt_dir = self._le_gt.text()
        img_dir = self._le_eval_img.text()
        if not os.path.isdir(gt_dir) or not os.path.isdir(img_dir):
            QMessageBox.warning(self, "알림", "GT 라벨 폴더와 이미지 폴더를 선택하세요.")
            return
        n = self._eval_model_list.rowCount()
        if n == 0:
            QMessageBox.warning(self, "알림", "평가할 모델을 추가하세요.")
            return

        entries = []
        for i in range(n):
            path = self._eval_model_list.item(i, 0).text()
            combo = self._eval_model_list.cellWidget(i, 1)
            mtype = combo.currentData() or "yolo"
            entries.append((path, mtype))

        self._current_task = self._combo_task.currentText()

        if self._current_task == "Detection":
            # GT 클래스 스캔
            gt_classes = self._scan_gt_classes(gt_dir, img_dir)
            # 모델별 클래스 정보 수집
            model_infos = []
            for path, mtype in entries:
                mname = os.path.basename(path)
                try:
                    mi = load_model(path, model_type=mtype)
                except Exception as exc:
                    QMessageBox.critical(self, "오류", f"{mname} 로드 실패: {exc}")
                    return
                if mtype == "darknet" and mi.names:
                    # darknet: convert_darknet_to_unified 후 출력 클래스 = UNIFIED_NAMES
                    mcls = sorted(UNIFIED_NAMES.items())
                elif mi.names:
                    mcls = sorted(mi.names.items())
                else:
                    mcls = []  # 이름 없음 → 다이얼로그에서 사용자 지정
                model_infos.append((mname, mtype, mcls))

            # 매핑 다이얼로그
            from ui.class_mapping_dialog import ClassMappingDialog
            dlg = ClassMappingDialog(gt_classes, model_infos, self,
                                     prev_mappings=self._last_mappings,
                                     prev_mapped_only=self._last_mapped_only)
            if dlg.exec() != QDialog.Accepted:
                return
            per_model_mappings, mapped_only = dlg.get_result()
            self._last_mappings = per_model_mappings
            self._last_mapped_only = mapped_only
        else:
            per_model_mappings, mapped_only = {}, True

        self._table.setRowCount(0)
        self._lbl_summary.clear()
        self._btn_eval.setEnabled(False)
        self._btn_eval_stop.setEnabled(True)
        self._eval_prog.setValue(0)
        self._eval_row = 0
        self._eval_results = []

        if self._current_task == "Classification":
            self._eval_worker = _ClsEvalWorker(entries, img_dir, gt_dir)
        else:
            classmap = _parse_classmap(self._te_classmap.toPlainText())
            self._eval_worker = _EvalWorker(
                entries, img_dir, gt_dir, self._spin_eval_conf.value(), classmap,
                per_model_mappings=per_model_mappings, mapped_only=mapped_only,
            )

        self._eval_worker.progress.connect(self._on_eval_progress)
        self._eval_worker.model_done.connect(self._on_model_done)
        self._eval_worker.all_done.connect(self._on_eval_all_done)
        self._eval_worker.error.connect(lambda e: QMessageBox.critical(self, "오류", e))
        self._eval_worker.start()

    def _scan_gt_classes(self, gt_dir, img_dir):
        """GT 라벨에서 고유 클래스 ID 스캔 → [(id, name), ...]"""
        class_ids = set()
        gt_data = _load_yolo_labels(gt_dir, img_dir, ignore_classes=set())
        for boxes in gt_data.values():
            for b in boxes:
                class_ids.add(b[0])
        # classmap 텍스트에서 이름 가져오기
        classmap_names = {}
        for line in self._te_classmap.toPlainText().strip().splitlines():
            if ":" in line:
                try:
                    k, v = line.split(":", 1)
                    classmap_names[int(k.strip())] = v.strip()
                except ValueError:
                    pass
        return sorted((cid, classmap_names.get(cid, UNIFIED_NAMES.get(cid, f"class_{cid}")))
                      for cid in class_ids)

    def _on_eval_progress(self, name, cur, total):
        self._eval_prog.setMaximum(100)
        pct = int(cur / max(total, 1) * 100)
        self._eval_prog.setValue(pct)
        self._eval_prog.setFormat(f"{name}: {cur}/{total} ({pct}%)")

    def _on_model_done(self, name, metrics):
        row = self._eval_row
        self._table.insertRow(row)
        is_cls = getattr(self, '_current_task', 'Detection') == 'Classification'

        def _fmt(v):
            """값을 00.0000% 형식으로 포맷"""
            return f"{v * 100:.4f}%"

        if is_cls:
            self._table.setItem(row, 0, QTableWidgetItem(name))
            self._table.setItem(row, 1, QTableWidgetItem(_fmt(metrics.get('accuracy', 0))))
            self._table.setItem(row, 2, QTableWidgetItem(_fmt(metrics['precision'])))
            self._table.setItem(row, 3, QTableWidgetItem(_fmt(metrics['recall'])))
            self._table.setItem(row, 4, QTableWidgetItem(_fmt(metrics['f1'])))
            btn = QPushButton("상세보기")
            detail = metrics.get("detail", {})
            btn.clicked.connect(lambda _, n=name, d=detail: self._show_cls_detail(n, d))
            self._table.setCellWidget(row, 5, btn)
            score_col = 1
        else:
            self._table.setItem(row, 0, QTableWidgetItem(name))
            self._table.setItem(row, 1, QTableWidgetItem(_fmt(metrics['mAP50'])))
            self._table.setItem(row, 2, QTableWidgetItem(_fmt(metrics['mAP50_95'])))
            self._table.setItem(row, 3, QTableWidgetItem(_fmt(metrics['precision'])))
            self._table.setItem(row, 4, QTableWidgetItem(_fmt(metrics['recall'])))
            self._table.setItem(row, 5, QTableWidgetItem(_fmt(metrics['f1'])))
            btn = QPushButton("상세보기")
            detail = metrics.get("detail", {})
            btn.clicked.connect(lambda _, n=name, d=detail: self._show_detail(n, d))
            self._table.setCellWidget(row, 6, btn)
            score_col = 1

        # 최고 점수 하이라이트
        best_row, best_val = 0, -1.0
        for r in range(self._table.rowCount()):
            item = self._table.item(r, score_col)
            if item:
                try:
                    v = float(item.text().replace('%', ''))
                except ValueError:
                    v = 0.0
                if v > best_val:
                    best_val, best_row = v, r
        from PySide6.QtGui import QColor
        ncols = self._table.columnCount() - 1
        for r in range(self._table.rowCount()):
            is_best = (r == best_row)
            bg_hex, fg_hex = theme.best_cell_colors() if is_best else theme.normal_cell_colors()
            bg = QColor(bg_hex)
            fg = QColor(fg_hex)
            for c in range(ncols):
                item = self._table.item(r, c)
                if item:
                    item.setBackground(bg)
                    item.setForeground(fg)

        self._eval_results.append((name, metrics))
        self._eval_row += 1

    def _show_detail(self, model_name, detail):
        from PySide6.QtWidgets import QDialog
        dlg = QDialog(self)
        dlg.setWindowTitle(f"클래스별 상세 — {model_name}")
        dlg.resize(600, 400)
        lay = QVBoxLayout(dlg)
        tbl = QTableWidget()
        class_keys = sorted(k for k in detail if k != "__overall__")
        tbl.setColumnCount(8)
        tbl.setHorizontalHeaderLabels(["클래스", "AP@50", "Precision", "Recall", "F1", "TP", "FP", "FN"])
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tbl.setRowCount(len(class_keys))
        for r, cid in enumerate(class_keys):
            v = detail[cid]
            tbl.setItem(r, 0, QTableWidgetItem(UNIFIED_NAMES.get(cid, str(cid))))
            tbl.setItem(r, 1, QTableWidgetItem(f"{v['ap']*100:.4f}%"))
            tbl.setItem(r, 2, QTableWidgetItem(f"{v['precision']*100:.4f}%"))
            tbl.setItem(r, 3, QTableWidgetItem(f"{v['recall']*100:.4f}%"))
            tbl.setItem(r, 4, QTableWidgetItem(f"{v['f1']*100:.4f}%"))
            tbl.setItem(r, 5, QTableWidgetItem(str(v['tp'])))
            tbl.setItem(r, 6, QTableWidgetItem(str(v['fp'])))
            tbl.setItem(r, 7, QTableWidgetItem(str(v['fn'])))
        lay.addWidget(tbl)
        dlg.exec()

    def _show_cls_detail(self, model_name, detail):
        from PySide6.QtWidgets import QDialog
        dlg = QDialog(self)
        dlg.setWindowTitle(f"클래스별 상세 — {model_name}")
        dlg.resize(600, 400)
        lay = QVBoxLayout(dlg)
        tbl = QTableWidget()
        class_keys = sorted(k for k in detail if k != "__overall__")
        tbl.setColumnCount(6)
        tbl.setHorizontalHeaderLabels(["클래스", "Precision", "Recall", "F1", "TP", "FP"])
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tbl.setRowCount(len(class_keys))
        for r, cid in enumerate(class_keys):
            v = detail[cid]
            tbl.setItem(r, 0, QTableWidgetItem(str(cid)))
            tbl.setItem(r, 1, QTableWidgetItem(f"{v['precision']*100:.4f}%"))
            tbl.setItem(r, 2, QTableWidgetItem(f"{v['recall']*100:.4f}%"))
            tbl.setItem(r, 3, QTableWidgetItem(f"{v['f1']*100:.4f}%"))
            tbl.setItem(r, 4, QTableWidgetItem(str(v['tp'])))
            tbl.setItem(r, 5, QTableWidgetItem(str(v['fp'])))
        lay.addWidget(tbl)
        dlg.exec()

    def _on_eval_all_done(self):
        self._btn_eval.setEnabled(True)
        self._btn_eval_stop.setEnabled(False)
        self._eval_prog.setValue(100)

    def _stop_eval(self):
        if hasattr(self, '_eval_worker') and self._eval_worker.isRunning():
            self._eval_worker.stop()

    def _export_excel(self):
        if not hasattr(self, '_eval_results') or not self._eval_results:
            QMessageBox.warning(self, "알림", "내보낼 결과가 없습니다.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Excel 저장", "eval_results.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()

            # 요약 시트
            ws = wb.active
            ws.title = "Summary"
            hdr = ["모델", "mAP@50", "mAP@50:95", "Precision", "Recall", "F1"]
            bold = Font(bold=True)
            for c, h in enumerate(hdr, 1):
                ws.cell(1, c, h).font = bold
            for r, (name, m) in enumerate(self._eval_results, 2):
                ws.cell(r, 1, name)
                ws.cell(r, 2, float(m["mAP50"]))
                ws.cell(r, 3, float(m["mAP50_95"]))
                ws.cell(r, 4, float(m["precision"]))
                ws.cell(r, 5, float(m["recall"]))
                ws.cell(r, 6, float(m["f1"]))

            # 모델별 상세 시트
            for name, m in self._eval_results:
                detail = m.get("detail", {})
                import re
                sname = os.path.splitext(name)[0]
                sname = re.sub(r'[\\/*?:\[\]]', '_', sname)[:31]
                # 중복 시트명 방지
                existing = [s.title for s in wb.worksheets]
                base, idx = sname, 1
                while sname in existing:
                    sname = f"{base[:28]}_{idx}"
                    idx += 1
                ds = wb.create_sheet(sname)
                dhdr = ["클래스", "AP@50", "Precision", "Recall", "F1", "TP", "FP", "FN"]
                for c, h in enumerate(dhdr, 1):
                    ds.cell(1, c, h).font = bold
                row = 2
                for cid in sorted(k for k in detail if k != "__overall__"):
                    v = detail[cid]
                    ds.cell(row, 1, UNIFIED_NAMES.get(cid, str(cid)))
                    ds.cell(row, 2, float(v["ap"]))
                    ds.cell(row, 3, float(v["precision"]))
                    ds.cell(row, 4, float(v["recall"]))
                    ds.cell(row, 5, float(v["f1"]))
                    ds.cell(row, 6, int(v["tp"]))
                    ds.cell(row, 7, int(v["fp"]))
                    ds.cell(row, 8, int(v["fn"]))
                    row += 1

            wb.save(path)
            QMessageBox.information(self, "완료", f"저장 완료\n{path}")
        except ImportError:
            QMessageBox.critical(self, "오류", "openpyxl이 설치되어 있지 않습니다.\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "오류", str(e))

    def _export_html(self):
        if not hasattr(self, '_eval_results') or not self._eval_results:
            QMessageBox.warning(self, "알림", "내보낼 결과가 없습니다.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "HTML 저장", "eval_report.html", "HTML (*.html)")
        if not path:
            return
        try:
            is_cls = getattr(self, '_current_task', 'Detection') == 'Classification'
            html = _generate_html_report(self._eval_results, is_cls)
            with open(path, "w", encoding="utf-8") as f:
                f.write(html)
            QMessageBox.information(self, "완료", f"HTML 리포트 저장 완료\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "오류", str(e))


def _generate_html_report(eval_results, is_cls=False):
    """평가 결과를 HTML 리포트로 생성 (인라인 Chart.js)"""
    import json, datetime

    # 요약 테이블 데이터
    if is_cls:
        cols = ["모델", "Accuracy", "Precision", "Recall", "F1"]
        rows = []
        for name, m in eval_results:
            rows.append([name, f"{m.get('accuracy',0)*100:.4f}%", f"{m['precision']*100:.4f}%",
                         f"{m['recall']*100:.4f}%", f"{m['f1']*100:.4f}%"])
    else:
        cols = ["모델", "mAP@50", "mAP@50:95", "Precision", "Recall", "F1"]
        rows = []
        for name, m in eval_results:
            rows.append([name, f"{m['mAP50']*100:.4f}%", f"{m['mAP50_95']*100:.4f}%",
                         f"{m['precision']*100:.4f}%", f"{m['recall']*100:.4f}%", f"{m['f1']*100:.4f}%"])

    # 차트 데이터
    model_names = [r[0] for r in rows]
    if is_cls:
        chart_labels = ["Accuracy", "Precision", "Recall", "F1"]
        chart_data = {l: [float(rows[i][j+1]) for i in range(len(rows))] for j, l in enumerate(chart_labels)}
    else:
        chart_labels = ["mAP@50", "mAP@50:95", "Precision", "Recall", "F1"]
        chart_data = {l: [float(rows[i][j+1]) for i in range(len(rows))] for j, l in enumerate(chart_labels)}

    colors = ["#4CAF50", "#2196F3", "#FF9800", "#9C27B0", "#F44336", "#00BCD4"]

    datasets_js = []
    for i, label in enumerate(chart_labels):
        datasets_js.append({
            "label": label,
            "data": chart_data[label],
            "backgroundColor": colors[i % len(colors)] + "88",
            "borderColor": colors[i % len(colors)],
            "borderWidth": 1,
        })

    # 상세 테이블 (모델별)
    detail_sections = ""
    for name, m in eval_results:
        detail = m.get("detail", {})
        class_keys = sorted(k for k in detail if k != "__overall__")
        if not class_keys:
            continue
        detail_sections += f'<h3>{name} — 클래스별 상세</h3><table><tr>'
        if is_cls:
            detail_sections += '<th>클래스</th><th>Precision</th><th>Recall</th><th>F1</th></tr>'
            for cid in class_keys:
                v = detail[cid]
                detail_sections += (f'<tr><td>{cid}</td><td>{v["precision"]*100:.4f}%</td>'
                                    f'<td>{v["recall"]*100:.4f}%</td><td>{v["f1"]*100:.4f}%</td></tr>')
        else:
            detail_sections += '<th>클래스</th><th>AP@50</th><th>Precision</th><th>Recall</th><th>F1</th><th>TP</th><th>FP</th><th>FN</th></tr>'
            for cid in class_keys:
                v = detail[cid]
                cname = str(cid)
                detail_sections += (f'<tr><td>{cname}</td><td>{v["ap"]*100:.4f}%</td>'
                                    f'<td>{v["precision"]*100:.4f}%</td><td>{v["recall"]*100:.4f}%</td>'
                                    f'<td>{v["f1"]*100:.4f}%</td><td>{v["tp"]}</td>'
                                    f'<td>{v["fp"]}</td><td>{v["fn"]}</td></tr>')
        detail_sections += '</table>'

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    task_name = "Classification" if is_cls else "Detection"

    return f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<title>평가 리포트 — {task_name}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
body {{ font-family: 'Segoe UI', sans-serif; margin: 40px; background: #f5f5f5; color: #333; }}
h1 {{ color: #1a237e; }} h2 {{ color: #283593; margin-top: 30px; }} h3 {{ color: #3949ab; }}
table {{ border-collapse: collapse; width: 100%; margin: 10px 0; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
th, td {{ border: 1px solid #ddd; padding: 8px 12px; text-align: center; }}
th {{ background: #e8eaf6; font-weight: 600; }}
tr:nth-child(even) {{ background: #fafafa; }}
.chart-container {{ max-width: 800px; margin: 20px auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
.meta {{ color: #666; font-size: 0.9em; }}
</style></head><body>
<h1>모델 평가 리포트</h1>
<p class="meta">태스크: {task_name} | 생성: {now} | 모델 수: {len(eval_results)}</p>

<h2>요약</h2>
<table><tr>{''.join(f'<th>{c}</th>' for c in cols)}</tr>
{''.join('<tr>' + ''.join(f'<td>{c}</td>' for c in row) + '</tr>' for row in rows)}
</table>

<h2>비교 차트</h2>
<div class="chart-container"><canvas id="chart"></canvas></div>
<script>
new Chart(document.getElementById('chart'), {{
  type: 'bar',
  data: {{ labels: {json.dumps(model_names)}, datasets: {json.dumps(datasets_js)} }},
  options: {{ responsive: true, scales: {{ y: {{ beginAtZero: true, max: 1.0 }} }} }}
}});
</script>

<h2>클래스별 상세</h2>
{detail_sections}

</body></html>"""
